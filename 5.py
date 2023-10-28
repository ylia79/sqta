import openpyxl 

  
header = ["Student Name", "Subject1", "Subject2", "Subject3", "Total"] 

sname = ["Carls", "James", "Paul", "Philip", "Smith", "Thomson", "Rhodey", "Stark", "Gary", "AnneMarie"] 


marks = [50, 45, 60, 55, 70, 45, 67, 78, 89, 90, 30] 


workbook = openpyxl.Workbook() 

sheet = workbook.active 

sheet.title = "Report" 

for col_num, value in enumerate(header, 1): 

    sheet.cell(row=1, column=col_num, value=value) 

for row_num, name in enumerate(sname, 2): 

    sheet.cell(row=row_num, column=1, value=name) 

for row_num, mark in enumerate(marks, 2): 

    for col_num in range(2, 5): 

        sheet.cell(row=row_num, column=col_num, value=mark) 

for row_num in range(2, len(sname) + 2): 
    total = sum(marks[row_num - 2:row_num + 1]) 

    sheet.cell(row=row_num, column=5, value=total) 

workbook.save("student.xlsx") 

print("Excel File Created!") 