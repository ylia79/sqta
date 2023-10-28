import openpyxl 

class ExcelReader: 

    def __init__(self, input_file): 

        self.input_file = input_file 

    def read(self): 

        workbook = openpyxl.load_workbook(self.input_file, data_only=True) 

        sheet = workbook.active 

        flag = False 

        count = 0 

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=sheet.max_column - 1): 

            for cell in row: 

                if cell.value is not None and isinstance(cell.value, (int, float)): 

                    if cell.value >= 60: 

                        flag = True 

                        if flag: 

                            count += 1 

                            flag = False 

                        break 
        print("Total number of students who scored more than 60 in one or more subjects:", count) 
if __name__ == "__main__": 
    excel_reader = ExcelReader("student.xlsx") 
    excel_reader.read() 